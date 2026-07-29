"""Bounded CSV/XLSX ingestion that preserves immutable source lineage."""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal
from hashlib import sha256
from io import BytesIO, StringIO
import json
from types import MappingProxyType
from typing import Iterable, Mapping
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.domain.transactions import RawRecord
from app.services.normalization import ColumnMapping, NormalizationResult, normalize_record


MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_SHEETS = 20
MAX_ROWS = 10_000
MAX_COLUMNS = 200
HEADER_GUESS_ROWS = 10
MAX_ARCHIVE_MEMBERS = 2_000
MAX_ARCHIVE_MEMBER_BYTES = 32 * 1024 * 1024
MAX_ARCHIVE_EXPANDED_BYTES = 64 * 1024 * 1024
MAX_ARCHIVE_COMPRESSION_RATIO = 200


class IngestionError(ValueError):
    """Base error for a source file that cannot be safely ingested."""


class UnsupportedFileError(IngestionError):
    """The input does not use a supported, macro-free file format."""


class IngestionLimitError(IngestionError):
    """The input exceeds an explicit safety limit."""


@dataclass(frozen=True, slots=True)
class IngestionMapping:
    """Source layout settings wrapped around the shared normalization mapping."""

    columns: ColumnMapping
    sheet_name: str | None = None
    header_row: int = 1
    source_file_column: str | None = None
    max_rows: int = MAX_ROWS
    max_columns: int = MAX_COLUMNS

    def __post_init__(self) -> None:
        if self.header_row < 1:
            raise ValueError("header_row must be positive")
        if self.max_rows < 1:
            raise ValueError("max_rows must be positive")
        if self.max_columns < 1:
            raise ValueError("max_columns must be positive")


@dataclass(frozen=True, slots=True)
class HeaderGuess:
    row_number: int
    values: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class WorkbookInspection:
    sheet_names: tuple[str, ...]
    header_guesses: Mapping[str, tuple[HeaderGuess, ...]]

    def __post_init__(self) -> None:
        object.__setattr__(self, "header_guesses", MappingProxyType(dict(self.header_guesses)))


# Kept as a compatibility name for callers written against the implementation plan.
IngestionPreview = WorkbookInspection


@dataclass(frozen=True, slots=True)
class IngestionBatch:
    source_filename: str
    source_file_sha256: str
    raw_records: tuple[RawRecord, ...]
    normalization_results: tuple[NormalizationResult, ...]
    counts: Mapping[str, int]

    def __post_init__(self) -> None:
        object.__setattr__(self, "counts", MappingProxyType(dict(self.counts)))


def inspect_workbook(data: bytes) -> WorkbookInspection:
    """Return only bounded workbook structure for mapping selection, never row data."""
    _validate_data_size(data)
    workbook = _load_xlsx(data)
    try:
        sheet_names = tuple(workbook.sheetnames)
        if len(sheet_names) > MAX_SHEETS:
            raise IngestionLimitError("workbook sheet limit exceeded")
        guesses: dict[str, tuple[HeaderGuess, ...]] = {}
        for worksheet in workbook.worksheets:
            if worksheet.max_column > MAX_COLUMNS:
                raise IngestionLimitError("workbook column limit exceeded")
            candidates: list[HeaderGuess] = []
            for row_number, values in enumerate(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=min(worksheet.max_row, HEADER_GUESS_ROWS),
                    values_only=True,
                ),
                start=1,
            ):
                text_values = tuple("" if value is None else str(value) for value in values)
                if any(text_values):
                    candidates.append(HeaderGuess(row_number=row_number, values=text_values))
            guesses[worksheet.title] = tuple(
                sorted(candidates, key=lambda guess: (-sum(bool(value) for value in guess.values), guess.row_number))
            )
        return WorkbookInspection(sheet_names=sheet_names, header_guesses=guesses)
    finally:
        workbook.close()


def ingest_rows(
    data: bytes, filename: str, mapping: ColumnMapping | IngestionMapping
) -> IngestionBatch:
    """Create immutable raw rows and normalization results from one CSV or XLSX upload."""
    _validate_data_size(data)
    configured_mapping = _coerce_mapping(mapping)
    suffix = _file_suffix(filename)
    if suffix == ".csv":
        rows = _read_csv(data, configured_mapping)
    elif suffix == ".xlsx":
        rows = _read_xlsx(data, configured_mapping)
    elif suffix == ".xlsm":
        raise UnsupportedFileError("macro-enabled workbooks are not supported")
    else:
        raise UnsupportedFileError("only .csv and .xlsx files are supported")

    source_file_sha256 = sha256(data).hexdigest()
    ingested_at = datetime.now(timezone.utc)
    raw_records: list[RawRecord] = []
    normalization_results: list[NormalizationResult] = []
    for source_sheet, source_row_number, values in rows:
        row_sha256 = _row_sha256(values)
        raw = RawRecord(
            id=_raw_record_id(source_file_sha256, source_sheet, source_row_number, row_sha256),
            source_filename=filename,
            source_file_sha256=source_file_sha256,
            source_sheet=source_sheet,
            source_row_number=source_row_number,
            raw_values=values,
            raw_row_sha256=row_sha256,
            ingested_at=ingested_at,
        )
        raw_records.append(raw)
        normalization_results.append(normalize_record(raw, configured_mapping.columns))

    normalized_count = sum(result.transaction is not None for result in normalization_results)
    return IngestionBatch(
        source_filename=filename,
        source_file_sha256=source_file_sha256,
        raw_records=tuple(raw_records),
        normalization_results=tuple(normalization_results),
        counts={
            "raw_records": len(raw_records),
            "normalized": normalized_count,
            "quarantined": len(raw_records) - normalized_count,
        },
    )


def _coerce_mapping(mapping: ColumnMapping | IngestionMapping) -> IngestionMapping:
    if isinstance(mapping, IngestionMapping):
        return mapping
    if isinstance(mapping, ColumnMapping):
        return IngestionMapping(columns=mapping)
    raise TypeError("mapping must be a ColumnMapping or IngestionMapping")


def _validate_data_size(data: bytes) -> None:
    if not data:
        raise IngestionLimitError("source file is empty")
    if len(data) > MAX_FILE_BYTES:
        raise IngestionLimitError("source file size limit exceeded")


def _file_suffix(filename: str) -> str:
    if not filename or "." not in filename:
        raise UnsupportedFileError("source filename must have an extension")
    return "." + filename.rsplit(".", 1)[1].lower()


def _read_csv(data: bytes, mapping: IngestionMapping) -> Iterable[tuple[str, int, Mapping[str, object]]]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise UnsupportedFileError("CSV must be UTF-8 encoded") from error
    reader = csv.reader(StringIO(text, newline=""))
    header = _read_header_row(reader, mapping.header_row)
    _validate_headers(header, mapping)
    rows: list[tuple[str, int, Mapping[str, object]]] = []
    for row_number, values in enumerate(reader, start=mapping.header_row + 1):
        if not any(value.strip() for value in values):
            continue
        if len(values) > mapping.max_columns:
            raise IngestionLimitError("CSV column limit exceeded")
        if len(values) > len(header):
            raise IngestionError("CSV row has more columns than its header")
        rows.append(("CSV", row_number, dict(zip(header, values, strict=False))))
        _check_row_limit(rows, mapping.max_rows)
    return rows


def _read_header_row(reader: csv.reader, header_row: int) -> tuple[str, ...]:
    for row_number, row in enumerate(reader, start=1):
        if row_number == header_row:
            return _header_values(row)
    raise IngestionError("configured header row does not exist")


def _read_xlsx(data: bytes, mapping: IngestionMapping) -> Iterable[tuple[str, int, Mapping[str, object]]]:
    workbook = _load_xlsx(data)
    try:
        if len(workbook.sheetnames) > MAX_SHEETS:
            raise IngestionLimitError("workbook sheet limit exceeded")
        worksheet = _select_worksheet(workbook, mapping.sheet_name)
        if worksheet.max_column > mapping.max_columns:
            raise IngestionLimitError("workbook column limit exceeded")
        header_values = next(
            worksheet.iter_rows(min_row=mapping.header_row, max_row=mapping.header_row, values_only=True),
            None,
        )
        if header_values is None:
            raise IngestionError("configured header row does not exist")
        header = _header_values(header_values)
        _validate_headers(header, mapping)
        rows: list[tuple[str, int, Mapping[str, object]]] = []
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=mapping.header_row + 1, values_only=True),
            start=mapping.header_row + 1,
        ):
            if not any(value is not None and str(value).strip() for value in values):
                continue
            row = dict(zip(header, values, strict=False))
            amount_value = row.get(mapping.columns.amount)
            if isinstance(amount_value, float):
                # openpyxl exposes numeric spreadsheet cells as binary floats. Convert
                # only the money adapter boundary; the domain continues to reject floats.
                row[mapping.columns.amount] = Decimal(str(amount_value))
            rows.append((worksheet.title, row_number, row))
            _check_row_limit(rows, mapping.max_rows)
        return rows
    finally:
        workbook.close()


def _load_xlsx(data: bytes):
    try:
        with ZipFile(BytesIO(data)) as archive:
            _validate_xlsx_archive(archive)
            if any(name.lower().endswith("vbaproject.bin") for name in archive.namelist()):
                raise UnsupportedFileError("macro-enabled workbooks are not supported")
        return load_workbook(BytesIO(data), read_only=True, data_only=False, keep_vba=False)
    except BadZipFile as error:
        raise UnsupportedFileError("input is not a valid XLSX workbook") from error
    except InvalidFileException as error:
        raise UnsupportedFileError("input is not a valid XLSX workbook") from error


def _validate_xlsx_archive(archive: ZipFile) -> None:
    members = archive.infolist()
    if len(members) > MAX_ARCHIVE_MEMBERS:
        raise IngestionLimitError("XLSX archive member count limit exceeded")
    total_expanded_bytes = 0
    for member in members:
        if member.file_size > MAX_ARCHIVE_MEMBER_BYTES:
            raise IngestionLimitError("XLSX archive member size limit exceeded")
        total_expanded_bytes += member.file_size
        if total_expanded_bytes > MAX_ARCHIVE_EXPANDED_BYTES:
            raise IngestionLimitError("XLSX archive total expanded size limit exceeded")
        if member.file_size == 0:
            continue
        if member.compress_size == 0:
            raise IngestionLimitError("XLSX archive compression ratio limit exceeded")
        if member.file_size / member.compress_size > MAX_ARCHIVE_COMPRESSION_RATIO:
            raise IngestionLimitError("XLSX archive compression ratio limit exceeded")


def _select_worksheet(workbook, sheet_name: str | None):
    if sheet_name is None:
        return workbook.active
    if sheet_name not in workbook.sheetnames:
        raise IngestionError("configured worksheet does not exist")
    return workbook[sheet_name]


def _header_values(values: Iterable[object]) -> tuple[str, ...]:
    header = tuple("" if value is None else str(value).strip() for value in values)
    if not header or not any(header):
        raise IngestionError("header row is empty")
    if any(not value for value in header):
        raise IngestionError("header row contains a blank column name")
    if len(set(header)) != len(header):
        raise IngestionError("header row contains duplicate column names")
    if len(header) > MAX_COLUMNS:
        raise IngestionLimitError("source column limit exceeded")
    return header


def _validate_headers(header: tuple[str, ...], mapping: IngestionMapping) -> None:
    required = (
        mapping.columns.transaction_id,
        mapping.columns.transaction_date,
        mapping.columns.posted_date,
        mapping.columns.description,
        mapping.columns.amount,
        mapping.columns.currency,
        mapping.columns.bank_account,
    )
    if mapping.source_file_column is not None:
        required += (mapping.source_file_column,)
    missing = tuple(column for column in required if column not in header)
    if missing:
        raise IngestionError("configured source columns are missing")


def _check_row_limit(rows: list[object], max_rows: int) -> None:
    if len(rows) > max_rows:
        raise IngestionLimitError("source row limit exceeded")


def _row_sha256(values: Mapping[str, object]) -> str:
    encoded = json.dumps(
        {key: _json_value(value) for key, value in values.items()},
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
    )
    return sha256(encoded.encode("utf-8")).hexdigest()


def _json_value(value: object) -> object:
    if isinstance(value, datetime):
        return {"type": "datetime", "value": value.isoformat()}
    if isinstance(value, date):
        return {"type": "date", "value": value.isoformat()}
    if isinstance(value, Decimal):
        return {"type": "decimal", "value": str(value)}
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return {"type": type(value).__name__, "value": str(value)}


def _raw_record_id(file_sha256: str, sheet: str, row_number: int, row_sha256: str) -> str:
    payload = {
        "file_sha256": file_sha256,
        "row_number": row_number,
        "row_sha256": row_sha256,
        "sheet": sheet,
    }
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return sha256(encoded.encode("utf-8")).hexdigest()
